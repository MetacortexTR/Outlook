from flask import Flask, render_template, request, send_file, jsonify, after_this_request, send_from_directory
import os
from find_emails import find_categorized_emails_in_file
import pandas as pd
from werkzeug.utils import secure_filename
import tempfile
import logging
from flask_cors import CORS
import shutil
import re
import csv
import traceback  # Hata izleme için ekledik
import time
from openpyxl.styles import Font
import sys

# Loglama ayarları
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('app.log', encoding='utf-8', mode='w')
    ]
)
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

app = Flask(__name__, static_folder='static')
CORS(app)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size

# Favicon için route
@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                             'favicon.ico', mimetype='image/vnd.microsoft.icon')

@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
    response.headers.add('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    return response

def is_csv_file(filename):
    return filename.lower().endswith(('.csv', '.CSV'))

def sanitize_filename(filename):
    # Dosya adından geçersiz karakterleri temizle
    filename = re.sub(r'[<>:"/\\|?*]', '', filename)
    # Boşlukları alt çizgi ile değiştir
    filename = filename.replace(' ', '_')
    return filename

def extract_company_name(email):
    # E-posta adresinden firma adını çıkar (@'den sonra, ilk nokta'ya kadar)
    try:
        domain = email.split('@')[1]  # @'den sonraki kısmı al
        company = domain.split('.')[0]  # ilk noktaya kadar olan kısmı al
        return company
    except:
        return ''

def extract_name_from_text(text):
    # "Ad Soyad <email@domain.com>" formatından ismi çıkar
    try:
        # İsim kısmını bul (< işaretinden önceki kısım)
        name = text.split('<')[0].strip()
        
        # Gereksiz ön ekleri temizle
        unwanted_prefixes = [
            'To:', 'Cc:', 'From:', 'Gönderen:', 'Gönderen :', 
            'To :', 'Cc :', 'From :', ';', 'Bilgi:', 'Bilgi :'
        ]
        
        for prefix in unwanted_prefixes:
            if name.startswith(prefix):
                name = name[len(prefix):].strip()
        
        return name
    except:
        return ''

def is_airtable_csv(file_path):
    """Airtable CSV dosyası olup olmadığını kontrol et (daha esnek)"""
    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            first_line = f.readline().strip().lower()
            # Farklı başlık varyasyonlarını kontrol et
            name_headers = ['name', 'ad', 'ad soyad']
            email_headers = ['email', 'mail', 'e-posta', 'e posta']
            return any(h in first_line for h in name_headers) and any(h in first_line for h in email_headers)
    except:
        return False

# Uploads klasörünü oluştur
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST', 'OPTIONS'])
def process_file():
    if request.method == 'OPTIONS':
        return '', 204
        
    temp_csv_path = None
    temp_excel_path = None
    start_time = time.time()
    
    try:
        logger.debug('Dosya yükleme isteği alındı')
        logger.debug(f'Request Files: {request.files}')
        logger.debug(f'Request Form: {request.form}')
        
        if 'file' not in request.files:
            logger.error('Dosya bulunamadı')
            return jsonify({'error': 'Dosya seçilmedi'}), 400
        
        file = request.files['file']
        excel_name = request.form.get('excel_name', 'email_listesi')
        excel_name = sanitize_filename(excel_name)
        
        if file.filename == '':
            logger.error('Dosya adı boş')
            return jsonify({'error': 'Dosya seçilmedi'}), 400
        
        if not is_csv_file(file.filename):
            logger.error(f'Geçersiz dosya formatı: {file.filename}')
            return jsonify({'error': 'Lütfen CSV dosyası yükleyin'}), 400
        
        # Geçici dosyalar için dizin oluştur
        temp_dir = tempfile.mkdtemp()
        temp_csv_path = os.path.join(temp_dir, 'input.csv')
        temp_excel_path = os.path.join(temp_dir, 'output.xlsx')
        
        logger.debug(f'Geçici dizin oluşturuldu: {temp_dir}')
        
        try:
            # CSV dosyasını kaydet
            file.save(temp_csv_path)
            file_size = os.path.getsize(temp_csv_path)
            logger.debug(f'Dosya geçici konuma kaydedildi. Boyut: {file_size} bytes')
            
            # Dosya içeriğini kontrol et ve BOM karakterini temizle
            with open(temp_csv_path, 'r', encoding='utf-8-sig') as f:
                content = f.read()
            
            with open(temp_csv_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            with open(temp_csv_path, 'r', encoding='utf-8') as f:
                first_few_lines = ''.join(f.readline() for _ in range(5))
                logger.debug(f'Dosya içeriği (ilk 5 satır):\n{first_few_lines}')
        except Exception as e:
            logger.error(f'Dosya kaydetme/okuma hatası: {str(e)}')
            logger.error(traceback.format_exc())
            return jsonify({'error': f'Dosya işlenirken hata oluştu: {str(e)}'}), 400
        
        # E-posta adreslerini bul
        logger.debug('E-posta adresleri aranıyor...')
        try:
            categorized_data = find_categorized_emails_in_file(temp_csv_path)
            total_emails = sum(len(data) for data in categorized_data.values())
            logger.debug(f'Bulunan toplam e-posta sayısı: {total_emails}')
            if categorized_data:
                for category, data in categorized_data.items():
                    logger.debug(f'{category}: {len(data)} adet e-posta bulundu')
        except Exception as e:
            logger.error(f'E-posta arama hatası: {str(e)}')
            logger.error(traceback.format_exc())
            return jsonify({'error': f'E-posta adresleri işlenirken hata oluştu: {str(e)}'}), 400
        
        if not categorized_data:
            logger.warning('Hiç e-posta adresi bulunamadı')
            return jsonify({'error': 'Hiç e-posta adresi bulunamadı'}), 400
        
        # Excel dosyası oluştur
        logger.debug('Excel dosyası oluşturuluyor...')
        
        try:
            # Tüm kategorilerdeki e-postaları ve ilgili bilgileri topla
            all_data = []
            seen_emails = set()  # Daha önce görülen e-postaları takip etmek için

            # Kategorileri sırayla işle (Kimden, Kime, Bilgi, Gizli sırasında)
            for category in ['Kimden', 'Kime', 'Bilgi', 'Gizli']:
                if category in categorized_data:
                    for data in categorized_data[category]:
                        email = data['email']
                        # Eğer bu e-posta daha önce görülmediyse ekle
                        if email not in seen_emails:
                            seen_emails.add(email)
                            original_text = data['original_text']
                            all_data.append({
                                'Kategori': category,
                                'E-posta Adresi': email,
                                'Firma Adı': extract_company_name(email),
                                'Ad Soyad': extract_name_from_text(original_text)
                            })
            
            # DataFrame oluştur
            df = pd.DataFrame(all_data)
            
            # Excel'e kaydet
            with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='E-posta Listesi')
                workbook = writer.book
                worksheet = writer.sheets['E-posta Listesi']
                
                # Türkçe karakter desteği için font ayarı
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.font = Font(name='Calibri')
                
                # Sütun genişliklerini ayarla
                worksheet.column_dimensions['A'].width = 20  # Kategori sütunu
                worksheet.column_dimensions['B'].width = 40  # E-posta Adresi sütunu
                worksheet.column_dimensions['C'].width = 30  # Firma Adı sütunu
                worksheet.column_dimensions['D'].width = 30  # Ad Soyad sütunu
            
            logger.debug(f'Excel dosyası oluşturuldu: {temp_excel_path}')
        except Exception as e:
            logger.error(f'Excel oluşturma hatası: {str(e)}')
            logger.error(traceback.format_exc())
            return jsonify({'error': f'Excel dosyası oluşturulurken hata oluştu: {str(e)}'}), 400
        
        # İşlem süresini hesapla
        process_time = time.time() - start_time
        logger.info(f'Toplam işlem süresi: {process_time:.2f} saniye')
        
        # Excel dosyasını gönder
        logger.debug('Excel dosyası gönderiliyor...')
        
        @after_this_request
        def cleanup(response):
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
                logger.debug('Geçici dizin silindi')
            except Exception as e:
                logger.error(f'Temizleme hatası: {str(e)}')
            return response
        
        return send_file(
            temp_excel_path,
            as_attachment=True,
            download_name=f'{excel_name}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f'Genel hata: {str(e)}')
        logger.error(traceback.format_exc())
        if temp_csv_path and os.path.exists(os.path.dirname(temp_csv_path)):
            shutil.rmtree(os.path.dirname(temp_csv_path), ignore_errors=True)
        return jsonify({'error': f'Hata oluştu: {str(e)}'}), 500

@app.route('/merge_excel', methods=['POST', 'OPTIONS'])
def merge_excel():
    if request.method == 'OPTIONS':
        return '', 204
        
    temp_dir = None
    temp_excel_path = None
    start_time = time.time()
    
    try:
        logger.debug('Excel birleştirme isteği alındı')
        
        if 'files' not in request.files:
            logger.error('Dosya bulunamadı')
            return jsonify({'error': 'Dosya seçilmedi'}), 400
            
        files = request.files.getlist('files')
        excel_name = request.form.get('excel_name', 'birlesik_liste')
        excel_name = sanitize_filename(excel_name)
        
        if not files:
            logger.error('Dosya seçilmedi')
            return jsonify({'error': 'Lütfen en az bir Excel dosyası seçin'}), 400
            
        # Geçici dizin oluştur
        temp_dir = tempfile.mkdtemp()
        temp_excel_path = os.path.join(temp_dir, 'merged_output.xlsx')
        
        # Tüm e-postaları topla
        all_emails = set()
        all_data = []
        
        for file in files:
            if file.filename == '':
                continue
                
            if not file.filename.lower().endswith('.xlsx'):
                logger.error(f'Geçersiz dosya formatı: {file.filename}')
                return jsonify({'error': 'Lütfen sadece Excel (.xlsx) dosyaları yükleyin'}), 400
                
            # Geçici dosya oluştur
            temp_file_path = os.path.join(temp_dir, secure_filename(file.filename))
            file.save(temp_file_path)
            
            try:
                # Excel dosyasını oku
                df = pd.read_excel(temp_file_path)
                
                # E-posta sütununu bul
                email_column = None
                for col in df.columns:
                    if 'e-posta' in col.lower() or 'mail' in col.lower() or 'email' in col.lower():
                        email_column = col
                        break
                
                if email_column is None:
                    logger.warning(f'{file.filename} dosyasında e-posta sütunu bulunamadı')
                    continue
                
                # Her satırı işle
                for _, row in df.iterrows():
                    email = str(row[email_column]).strip().lower()
                    if '@' in email and email not in all_emails:
                        all_emails.add(email)
                        # Diğer sütunları da ekle
                        data = {'E-posta Adresi': email}
                        for col in df.columns:
                            if col != email_column:
                                data[col] = row[col]
                        all_data.append(data)
                
            except Exception as e:
                logger.error(f'Excel okuma hatası ({file.filename}): {str(e)}')
                logger.error(traceback.format_exc())
                continue
        
        if not all_data:
            logger.warning('Hiç e-posta adresi bulunamadı')
            return jsonify({'error': 'Hiç e-posta adresi bulunamadı'}), 400
        
        # Yeni Excel dosyası oluştur
        df = pd.DataFrame(all_data)
        
        with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='E-posta Listesi')
            workbook = writer.book
            worksheet = writer.sheets['E-posta Listesi']
            
            # Türkçe karakter desteği için font ayarı
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = Font(name='Calibri')
            
            # Sütun genişliklerini ayarla
            for idx, col in enumerate(df.columns):
                worksheet.column_dimensions[chr(65 + idx)].width = 30
        
        # İşlem süresini hesapla
        process_time = time.time() - start_time
        logger.info(f'Toplam işlem süresi: {process_time:.2f} saniye')
        
        @after_this_request
        def cleanup(response):
            try:
                if temp_dir and os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    logger.debug('Geçici dizin silindi')
            except Exception as e:
                logger.error(f'Temizleme hatası: {str(e)}')
            return response
        
        return send_file(
            temp_excel_path,
            as_attachment=True,
            download_name=f'{excel_name}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f'Genel hata: {str(e)}')
        logger.error(traceback.format_exc())
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        return jsonify({'error': f'Hata oluştu: {str(e)}'}), 500

@app.route('/process_airtable', methods=['POST', 'OPTIONS'])
def process_airtable():
    if request.method == 'OPTIONS':
        return '', 204
        
    temp_csv_path = None
    temp_excel_path = None
    start_time = time.time()
    
    try:
        logger.debug('Airtable CSV işleme isteği alındı')
        
        if 'file' not in request.files:
            logger.error('Dosya bulunamadı')
            return jsonify({'error': 'Dosya seçilmedi'}), 400
        
        file = request.files['file']
        excel_name = request.form.get('excel_name', 'airtable_liste')
        excel_name = sanitize_filename(excel_name)
        
        if file.filename == '':
            logger.error('Dosya adı boş')
            return jsonify({'error': 'Dosya seçilmedi'}), 400
        
        if not is_csv_file(file.filename):
            logger.error(f'Geçersiz dosya formatı: {file.filename}')
            return jsonify({'error': 'Lütfen CSV dosyası yükleyin'}), 400
        
        # Geçici dosyalar için dizin oluştur
        temp_dir = tempfile.mkdtemp()
        temp_csv_path = os.path.join(temp_dir, 'input.csv')
        temp_excel_path = os.path.join(temp_dir, 'output.xlsx')
        
        logger.debug(f'Geçici dizin oluşturuldu: {temp_dir}')
        
        try:
            # CSV dosyasını kaydet
            file.save(temp_csv_path)
            file_size = os.path.getsize(temp_csv_path)
            logger.debug(f'Dosya geçici konuma kaydedildi. Boyut: {file_size} bytes')
            
            # Airtable CSV kontrolü
            if not is_airtable_csv(temp_csv_path):
                logger.error('Geçersiz Airtable CSV formatı')
                return jsonify({'error': 'Lütfen Airtable\'dan export edilmiş bir CSV dosyası yükleyin'}), 400
            
            # CSV'yi oku
            df = pd.read_csv(temp_csv_path, encoding='utf-8-sig')
            
            # Excel'e kaydet
            with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Airtable Verileri')
                workbook = writer.book
                worksheet = writer.sheets['Airtable Verileri']
                
                # Türkçe karakter desteği için font ayarı
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.font = Font(name='Calibri')
                
                # Sütun genişliklerini ayarla
                for idx, col in enumerate(df.columns):
                    worksheet.column_dimensions[chr(65 + idx)].width = 30
            
            logger.debug(f'Excel dosyası oluşturuldu: {temp_excel_path}')
            
        except Exception as e:
            logger.error(f'Dosya işleme hatası: {str(e)}')
            logger.error(traceback.format_exc())
            return jsonify({'error': f'Dosya işlenirken hata oluştu: {str(e)}'}), 400
        
        # İşlem süresini hesapla
        process_time = time.time() - start_time
        logger.info(f'Toplam işlem süresi: {process_time:.2f} saniye')
        
        @after_this_request
        def cleanup(response):
            try:
                if temp_dir and os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    logger.debug('Geçici dizin silindi')
            except Exception as e:
                logger.error(f'Temizleme hatası: {str(e)}')
            return response
        
        return send_file(
            temp_excel_path,
            as_attachment=True,
            download_name=f'{excel_name}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f'Genel hata: {str(e)}')
        logger.error(traceback.format_exc())
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        return jsonify({'error': f'Hata oluştu: {str(e)}'}), 500

@app.route('/compare_merge', methods=['POST', 'OPTIONS'])
def compare_merge():
    if request.method == 'OPTIONS':
        return '', 204

    temp_dir = None
    temp_excel_path = None
    start_time = time.time()

    try:
        logger.debug('Karşılaştır ve Birleştir isteği alındı')

        if 'file1' not in request.files or 'file2' not in request.files:
            logger.error('Dosya(lar) bulunamadı')
            return jsonify({'error': 'Lütfen iki Excel dosyası yükleyin'}), 400

        file1 = request.files['file1']
        file2 = request.files['file2']
        excel_name = request.form.get('excel_name', 'tum_kisiler')
        excel_name = sanitize_filename(excel_name)

        if file1.filename == '' or file2.filename == '':
            logger.error('Dosya adı boş')
            return jsonify({'error': 'Lütfen iki Excel dosyası yükleyin'}), 400

        if not file1.filename.lower().endswith('.xlsx') or not file2.filename.lower().endswith('.xlsx'):
            logger.error('Geçersiz dosya formatı')
            return jsonify({'error': 'Lütfen sadece .xlsx dosyaları yükleyin'}), 400

        # Geçici dizin oluştur
        temp_dir = tempfile.mkdtemp()
        temp_file1 = os.path.join(temp_dir, secure_filename(file1.filename))
        temp_file2 = os.path.join(temp_dir, secure_filename(file2.filename))
        temp_excel_path = os.path.join(temp_dir, 'output.xlsx')

        file1.save(temp_file1)
        file2.save(temp_file2)

        # Dosyaları oku
        df1 = pd.read_excel(temp_file1)
        df2 = pd.read_excel(temp_file2)

        # Excel'e iki ayrı sheet olarak kaydet
        with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as writer:
            df1.to_excel(writer, index=False, sheet_name='Birleştirilmiş')
            df2.to_excel(writer, index=False, sheet_name='Airtable')
            workbook = writer.book
            for sheet in writer.sheets.values():
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.font = Font(name='Calibri')
                # Sütun genişliklerini ayarla
                for idx, col in enumerate(sheet.iter_cols(1, sheet.max_column)):
                    sheet.column_dimensions[chr(65 + idx)].width = 30

        process_time = time.time() - start_time
        logger.info(f'Toplam işlem süresi: {process_time:.2f} saniye')

        @after_this_request
        def cleanup(response):
            try:
                if temp_dir and os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    logger.debug('Geçici dizin silindi')
            except Exception as e:
                logger.error(f'Temizleme hatası: {str(e)}')
            return response

        return send_file(
            temp_excel_path,
            as_attachment=True,
            download_name=f'{excel_name}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f'Genel hata: {str(e)}')
        logger.error(traceback.format_exc())
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        return jsonify({'error': f'Hata oluştu: {str(e)}'}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False) 